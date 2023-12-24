import sys

from openpyxl import load_workbook, worksheet, workbook
from openpyxl.styles import Alignment
import pyautogui
import datetime
import mss
import requests

def vid_info(file_name):  # 모듈화
    for i in range(len(file_name)):
        if (file_name[i] == 'I' and file_name[i + 1] == 'P' and file_name[i + 2] == 'C') or\
                (file_name[i] == 'C' and file_name[i+1] == 'H'):
            ch = file_name[i + 3]
        if file_name[i] == '2' and file_name[i + 1] == '0' and file_name[i + 2] == '2':
            day = file_name[i:i + 8]

    return ch, day


def excel_linecheck():
    wb = load_workbook(r"폐사체데이터.xlsx")
    ws = wb.worksheets[0]

    n = 2

    for i in range(n, 999999):
        if ws[f'E{i}'].value is not None:
            n += 1
        else:
            wb.save('./폐사체데이터.xlsx')
            break
    return n

def cell_absorb():
    wb = load_workbook(r'./폐사체데이터.xlsx')
    ws = wb.worksheets[0]

    ws.merge_cells(start_row=58, end_row=59, start_column=1, end_column=1)
    # ws['A39'].alignment = Alignment(vertical='center', horizontal='center')
    wb.save('./폐사체데이터.xlsx')


# cell_absorb()
# def chicdata_Save():
#     wb = load_workbook(r"폐사체데이터.xlsx")
#     ws = wb.worksheets[0]
#
#     ws[f'E{next}'].value = f'{self.xpos}, {self.ypos}'
#     ws[f'C{next}'].value = self.fileName
#     ws[f'A{next}'].value = int(day)
#     ws[f'B{next}'].value = int(ch)
#     ws[f'D{next}'].value = save_time

def find_download_pos(time, result, target):
    target = datetime.time(target[0], target[1], target[2])
    # print(time)
    for i in range(int(len(time) / 2)):
        atime = datetime.time.fromisoformat(time[0 + i])
        btime = datetime.time.fromisoformat(time[(int(len(time) / 2)) + i])
        if atime <= target <= btime:
            print('find')
            return i
            # print('find')
            # print(result)
            # print(len(result['text']))
            # print(len(result['left']))
            # print(len(result['top']))
            # print(result['text'])
            # print(result['left'])
            # pass
            # sys.exit()


        for j in range(len(result['text'])):
             if str(target) == result['text'][j]:
                print(result['left'][j])
                return

        # pyautogui.scroll(-320)

def post_message(token, channel, text):
    response = requests.post("https://slack.com/api/chat.postMessage",
                            headers={"Authorization": "Bearer " + token},
                            data={"channel": channel, "text": text}
                            )

def ch_decision(xywh):
    x = int(xywh[0][0])
    y = int(xywh[0][1])

    if x < 640:
        if y < 360:
            return 1
        elif y < 720:
            return 4
        else:
            return 7
    elif x < 1280:
        if y < 360:
            return 2
        elif y < 720:
            return 5
        else:
            return 8
    else:
        if y < 360:
            return 3
        elif y < 720:
            return 6
        else:
            9

def screenshot():
    # 스크린샷 저장
    output_filename = './Full_Img/screenshot.png'

    with mss.mss() as mss_instance:
        mss_instance.shot(output=output_filename)

def time_check():
    current_time = datetime.datetime.now()
    if (current_time.hour < 17) and (current_time.hour > 0):
        return True
    else:
        return False

def get_crop(x, y, crop_img, img_w, img_h):
    if x >= 4 and y >= 4:
        around_crop = crop_img[y - 4:y + 4, x - 4:x + 4]
        crop_y = y - 4
        crop_x = x - 4

        if x + 4 > img_w:
            w = img_w - crop_x
        else:
            w = 8

        if y + 4 > img_h:
            h = img_h - crop_y
        else:
            h = 8
    elif x >= 4 and y < 4:
        around_crop = crop_img[y - y:y + 4, x - 4:x + 4]
        crop_x = x - 4
        crop_y = 0
        if x + 4 > img_w:
            w = img_w - crop_x
        else:
            w = 8

        h = y + 4
    elif x < 4 and y >= 4:
        around_crop = crop_img[y - 4:y + 4, x - x:x + 4]
        crop_y = y - 4
        crop_x = 0
        w = x + 4

        if y + 4 > img_h:
            h = img_h - crop_y
        else:
            h = 8
    else:
        around_crop = crop_img[y - y:y + 4, x - x:x + 4]
        crop_x = 0
        crop_y = 0
        w = x + 4
        h = y + 4

    return crop_x, crop_y, w, h
