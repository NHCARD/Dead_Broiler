import datetime
import os
import sys

import mss
import openpyxl

from detect_main import main, parse_opt
import time
from func import time_check, screenshot

# def make_log:
#     time = datetime.datetime.now()
#     try:
#         wb = openpyxl.load_workbook(f'./log/{time.year}-{time.month}-{time.day}.xlsx')
#     except:
#         wb = openpyxl.workbook()


def predict():
    stime = datetime.datetime.now()

    while(True):
        # try:
        #     wb = openpyxl.load_workbook(f'./log/{stime.year}-{stime.month}-{stime.day}.xlsx')
        # except:
        #     wb = openpyxl.Workbook()
        screenshot()
        opt = parse_opt()
        opt.nosave = True
        opt.exist_ok = True
        # opt.log_workbook = wb
        opt.test = True
        opt.source = './Full_Img'
        # opt.source = r"C:\Users\inuri64\Downloads\실시간영상_샘플_AdobeExpress.mp4"
        # opt.view_img = True
        print(datetime.datetime.now())
        ch, current_time = main(opt)
        return ch, current_time

        time.sleep(15)
        # os.remove('./Full_Img/screenshot.png')

    print(f'time != 6 < x < 17 !  -- {datetime.datetime.now()}')
    time.sleep(300)


if __name__ == "__main__":
    predict()
