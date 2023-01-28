from PyQt5.QtCore import QDir, Qt, QUrl
from PyQt5.QtMultimedia import QMediaContent, QMediaPlayer, QMediaMetaData
from PyQt5.QtMultimediaWidgets import QVideoWidget
from PyQt5.QtWidgets import (QApplication, QFileDialog, QHBoxLayout, QLabel,
                             QPushButton, QSizePolicy, QSlider, QStyle, QVBoxLayout, QWidget, QInputDialog,
                             QMessageBox, QLineEdit)
from PyQt5.QtWidgets import QMainWindow, QWidget, QPushButton, QAction
from PyQt5.QtGui import QIcon
import sys
import os
import cv2
import numpy as np
from PIL import Image
from openpyxl import load_workbook, worksheet, workbook
from openpyxl.styles import Alignment

wb = load_workbook(r"폐사체데이터.xlsx")
ws = wb.worksheets[0]

next = 2

for i in range(next, 999999):
    if ws[f'E{i}'].value != None:
        next += 1
    else:
        break

class VideoPlayer(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("")

        self.mediaPlayer = QMediaPlayer(None, QMediaPlayer.VideoSurface)

        videoWidget = QVideoWidget()

        self.playButton = QPushButton()
        self.playButton.setEnabled(False)
        self.playButton.setIcon(self.style().standardIcon(QStyle.SP_MediaPlay))
        self.playButton.clicked.connect(self.play)

        self.positionSlider = QSlider(Qt.Horizontal)
        self.positionSlider.setRange(0, 0)
        self.positionSlider.sliderMoved.connect(self.setPosition)

        self.error = QLabel()
        self.error.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Maximum)

        openButton = QPushButton("Open Video")
        openButton.setToolTip("Open Video File")
        openButton.setStatusTip("Open Video File")
        openButton.setFixedHeight(24)
        openButton.clicked.connect(self.openFile)

        cap_btn = QPushButton('image')
        cap_btn.setFixedHeight(24)
        cap_btn.clicked.connect(self.cap_img)

        pbrate_label = QLineEdit()


        # Create a widget for window contents
        wid = QWidget(self)
        self.setCentralWidget(wid)

        # Create layouts to place inside widget
        controlLayout = QHBoxLayout()
        controlLayout.setContentsMargins(0, 0, 0, 0)
        controlLayout.addWidget(self.playButton)
        controlLayout.addWidget(self.positionSlider)

        layout = QVBoxLayout()
        layout.addWidget(videoWidget)
        layout.addLayout(controlLayout)
        layout.addWidget(self.error)
        layout.addWidget(openButton)
        layout.addWidget(cap_btn)

        # Set widget to contain window contents
        wid.setLayout(layout)

        self.mediaPlayer.setVideoOutput(videoWidget)
        self.mediaPlayer.stateChanged.connect(self.mediaStateChanged)
        self.mediaPlayer.positionChanged.connect(self.positionChanged)
        self.mediaPlayer.durationChanged.connect(self.durationChanged)
        self.mediaPlayer.error.connect(self.handleError)
        # self.mediaPlayer.

    def openFile(self):
        self.fileDir, _ = QFileDialog.getOpenFileName(self, "Open Movie",
                                                  QDir.homePath())

        print(self.fileDir)
        if self.fileDir != '':
            self.mediaPlayer.setMedia(
                QMediaContent(QUrl.fromLocalFile(self.fileDir)))
            self.playButton.setEnabled(True)

            for i in range(-1, -(int(len(self.fileDir))), -1):
                print(i)
                if self.fileDir[i] == '/':
                    self.fileName = self.fileDir[i+1:]
                    break
            print(self.fileName)

        print(self.mediaPlayer.metaData(QMediaMetaData.VideoFrameRate))

    def exitCall(self):
        sys.exit(app.exec_())

    def play(self):
        if self.mediaPlayer.state() == QMediaPlayer.PlayingState:
            self.mediaPlayer.pause()
            print(self.mediaPlayer.metaData(QMediaMetaData.VideoFrameRate))
            # print(self.mediaPlayer.)
        else:
            self.mediaPlayer.play()
            print(self.mediaPlayer.metaData(QMediaMetaData.Resolution))

    def mediaStateChanged(self, state):
        if self.mediaPlayer.state() == QMediaPlayer.PlayingState:
            self.playButton.setIcon(
                self.style().standardIcon(QStyle.SP_MediaPause))
        else:
            self.playButton.setIcon(
                self.style().standardIcon(QStyle.SP_MediaPlay))

    def positionChanged(self, position):
        self.positionSlider.setValue(position)
        print(self.positionSlider.value())

    def durationChanged(self, duration):
        self.positionSlider.setRange(0, duration)

    def setPosition(self, position):
        self.mediaPlayer.setPosition(position)

    def handleError(self):
        self.playButton.setEnabled(False)
        self.error.setText("Error: " + self.mediaPlayer.errorString())

    def rightKey_pressed(self):
        self.positionSlider.setValue(self.positionSlider.value() + 5000)
        print('asdf')
        # self.mediaPlayer.setPlaybackRate(5.0)

    def leftKey_pressed(self):
        self.positionSlider.setValue(self.positionSlider.value() - 5000)
        print('qwer')

    def vid_info(self, file_name):
        for i in range(len(file_name)):
            if file_name[i] == 'I' and file_name[i+1] == 'P' and file_name[i+2] == 'C':
                ch = file_name[i+3]
            if file_name[i] == '2' and file_name[i+1] == '0' and file_name[i+2] == '2':
                day = file_name[i:i+8]

        return ch, day

    def keyPressEvent(self, e):
        if e.key() == Qt.Key_Right:
            self.rightKey_pressed()
        elif e.key() == Qt.Key_Left:
            self.leftKey_pressed()

    def cap_img(self):
        global next
        sec = self.positionSlider.value() // 1000

        try:
            os.remove('original.png')
        except:
            pass

        os.system(fr"ffmpeg -vsync 2 -ss {sec} -t {sec+1} -i {'./0_8_IPC1_20220912083340.mp4'} -an -vf thumbnail=25 {'./original'}.png ")

        area, ok = QInputDialog.getText(self, 'input dialog', 'enter area : ')
        img = cv2.imread('original.png')
        pil_img = Image.fromarray(img)
        if ok and (area == '1' or area == '2' or area == '3' or area == '4'):
            reply = QMessageBox.question(self, 'Message', f'선택한 사분면이 {area}사분면이 맞습니까?',
                                                   QMessageBox.Yes | QMessageBox.No,
                                                   QMessageBox.Yes)
            if reply == QMessageBox.Yes:
                if area == '1':
                    cropped_img = pil_img.crop((0, 0, 1440, 810))
                    cropped_img = np.array(cropped_img)
                    x, y, _, _ = cv2.selectROI('click point', cropped_img, False)
                    if x == 0 and y == 0:
                        os.remove('original.png')
                        cv2.destroyWindow('click point')
                        return
                elif area == '2':
                    cropped_img = pil_img.crop((1441, 0, 2880, 1620))
                    cropped_img = np.array(cropped_img)
                    x, y, _, _ = cv2.selectROI('click point', cropped_img, False)
                    if x == 0 and y == 0:
                        os.remove('original.png')
                        cv2.destroyWindow('click point')
                        return
                    x += 1440
                elif area == '3':
                    cropped_img = pil_img.crop((0, 811, 1440, 1620))
                    cropped_img = np.array(cropped_img)
                    x, y, _, _ = cv2.selectROI('click point', cropped_img, False)
                    if x == 0 and y == 0:
                        os.remove('original.png')
                        cv2.destroyWindow('click point')
                        return
                    y += 810
                elif area == '4':
                    cropped_img = pil_img.crop((1441, 811, 2880, 1620))
                    cropped_img = np.array(cropped_img)
                    x, y, _, _ = cv2.selectROI('click point', cropped_img, False)
                    if x == 0 and y == 0:
                        os.remove('original.png')
                        cv2.destroyWindow('click point')
                        return
                    x += 1440
                    y += 810
                cv2.destroyWindow('click point')
                os.remove('original.png')
                print(x, y)
                ch, day = self.vid_info(self.fileName)
                save_time = str(sec//60) + ':' + str(sec - ((sec//60) * 60))
                # alignment = Alignment(vertical='center', horizontal='center')
                ws[f'E{next}'].value = f'{x}, {y}'
                ws[f'C{next}'].value = self.fileName
                ws[f'A{next}'].value = int(day)
                ws[f'B{next}'].value = int(ch)
                ws[f'D{next}'].value = save_time
                # ws[f'D{next}'].alignment = alignment
                next += 1
                wb.save('./폐사체데이터.xlsx')

app = QApplication(sys.argv)
videoplayer = VideoPlayer()
videoplayer.resize(640, 480)
videoplayer.show()
sys.exit(app.exec_())
