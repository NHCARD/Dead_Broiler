from openpyxl import load_workbook, worksheet, workbook
from openpyxl.styles import Alignment

def vid_info(file_name):  # 모듈화
    for i in range(len(file_name)):
        if (file_name[i] == 'I' and file_name[i + 1] == 'P' and file_name[i + 2] == 'C') or\
                (file_name[i] == 'C' and file_name[i+1] == 'H'):
            ch = file_name[i + 3]
        if file_name[i] == '2' and file_name[i + 1] == '0' and file_name[i + 2] == '2':
            day = file_name[i:i + 8]

    return ch, day


def excel_linecheck():
    wb = load_workbook(r"폐사체데이터.xlsx")
    ws = wb.worksheets[0]

    n = 2

    for i in range(n, 999999):
        if ws[f'E{i}'].value is not None:
            n += 1
        else:
            wb.save('./폐사체데이터.xlsx')
            break
    return n

def cell_absorb():
    wb = load_workbook(r'./폐사체데이터.xlsx')
    ws = wb.worksheets[0]

    ws.merge_cells(start_row=58, end_row=59, start_column=1, end_column=1)
    # ws['A39'].alignment = Alignment(vertical='center', horizontal='center')
    wb.save('./폐사체데이터.xlsx')


# cell_absorb()
# def chicdata_Save():
#     wb = load_workbook(r"폐사체데이터.xlsx")
#     ws = wb.worksheets[0]
#
#     ws[f'E{next}'].value = f'{self.xpos}, {self.ypos}'
#     ws[f'C{next}'].value = self.fileName
#     ws[f'A{next}'].value = int(day)
#     ws[f'B{next}'].value = int(ch)
#     ws[f'D{next}'].value = save_time