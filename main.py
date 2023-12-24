import cv2
from PyQt5.QtCore import QDir, Qt, QUrl
from PyQt5.QtMultimedia import QMediaContent, QMediaPlayer, QMediaMetaData
from PyQt5.QtMultimediaWidgets import QVideoWidget
from PyQt5.QtWidgets import (QApplication, QFileDialog, QHBoxLayout, QLabel,
                             QPushButton, QSizePolicy, QSlider, QStyle, QVBoxLayout, QWidget, QInputDialog,
                             QMessageBox, QLineEdit)
from PyQt5.QtWidgets import QMainWindow, QWidget, QPushButton, QAction
from PyQt5.QtGui import QIcon
import sys
import os
import numpy as np
from openpyxl import load_workbook, worksheet, workbook
from openpyxl.styles import Alignment
import matplotlib
matplotlib.use('Qt5Agg')
import matplotlib.pyplot as plt
from func import vid_info

class VideoPlayer(QMainWindow):
    def __init__(self, n = 0):
        super().__init__()
        self.next = n
        self.sec = None
        self.ypos = None
        self.xpos = None
        self.mode = None
        self.xlsx_name = '폐사체데이터_long.xlsx'
        self.setWindowTitle("")

        self.mediaPlayer = QMediaPlayer(None, QMediaPlayer.VideoSurface)

        videoWidget = QVideoWidget()

        self.playButton = QPushButton()
        self.playButton.setEnabled(False)
        self.playButton.setIcon(self.style().standardIcon(QStyle.SP_MediaPlay))
        self.playButton.clicked.connect(self.play)

        self.positionSlider = QSlider(Qt.Horizontal)
        self.positionSlider.setRange(0, 0)
        self.positionSlider.sliderMoved.connect(self.setPosition)
        self.positionSlider.valueChanged.connect(self.setPosition)

        self.error = QLabel()
        self.error.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Maximum)

        openButton = QPushButton("Open Video")
        openButton.setToolTip("Open Video File")
        openButton.setStatusTip("Open Video File")
        openButton.setFixedHeight(24)
        openButton.clicked.connect(self.openFile)

        cap_btn = QPushButton('image')
        cap_btn.setFixedHeight(24)
        cap_btn.clicked.connect(self.cap_img)

        openImage = QPushButton('Open Image')
        openImage.setFixedHeight(24)
        openImage.clicked.connect(self.open_image)

        self.pbrate_label = QPushButton('pbrate')
        self.pbrate_label.setFixedWidth(80)
        self.pbrate_label.clicked.connect(self.pbrate_change)


        # Create a widget for window contents
        wid = QWidget(self)
        self.setCentralWidget(wid)

        # Create layouts to place inside widget
        controlLayout = QHBoxLayout()
        controlLayout.setContentsMargins(0, 0, 0, 0)
        controlLayout.addWidget(self.playButton)
        controlLayout.addWidget(self.positionSlider)
        controlLayout.addWidget(self.pbrate_label)

        layout = QVBoxLayout()
        layout.addWidget(videoWidget)
        layout.addLayout(controlLayout)
        layout.addWidget(self.error)
        layout.addWidget(openButton)
        layout.addWidget(openImage)
        layout.addWidget(cap_btn)

        # Set widget to contain window contents
        wid.setLayout(layout)

        self.mediaPlayer = QMediaPlayer(self, flags=QMediaPlayer.VideoSurface)
        self.mediaPlayer.setVideoOutput(videoWidget)
        self.mediaPlayer.stateChanged.connect(self.mediaStateChanged)
        self.mediaPlayer.positionChanged.connect(self.positionChanged)
        self.mediaPlayer.durationChanged.connect(self.durationChanged)
        self.mediaPlayer.error.connect(self.handleError)
        # self.mediaPlayer.

    def openFile(self):
        '''
        파일 오픈
        '''
        self.fileDir, _ = QFileDialog.getOpenFileName(self, "Open Video", r'E:\4차\download\long_download\test\20231118', self.tr("Video Files (*.mp4)"))
        print(self.fileDir)
        self.mode = 0

        print(self.fileDir)
        if self.fileDir != '':
            self.mediaPlayer.setMedia(QMediaContent(QUrl.fromLocalFile(self.fileDir)))
            self.playButton.setEnabled(True)

            for i in range(-1, -(int(len(self.fileDir))), -1):
                print(i)
                if self.fileDir[i] == '/':
                    self.fileName = self.fileDir[i+1:]
                    break
            print(self.fileName)

    def exitCall(self):
        '''
        프로그램 종료
        '''
        sys.exit(app.exec_())

    def play(self):
        '''
        재생 상태 변경 함수
        '''
        if self.mediaPlayer.state() == QMediaPlayer.PlayingState:
            self.mediaPlayer.pause()
        else:
            self.mediaPlayer.play()

    def mediaStateChanged(self, state):
        '''
        재생상태 변경시
        :param state: 재생상태 나타냄
        '''
        if self.mediaPlayer.state() == QMediaPlayer.PlayingState:
            self.playButton.setIcon(
                self.style().standardIcon(QStyle.SP_MediaPause))
        else:
            self.playButton.setIcon(
                self.style().standardIcon(QStyle.SP_MediaPlay))

    def positionChanged(self, position):
        '''
        재생시간 막대 위치 이동
        :param position: 해당 값으로 이동
        '''
        self.positionSlider.setValue(position)

    def durationChanged(self, duration):
        '''
        재생시간 막대 크기 설정
        '''
        self.positionSlider.setRange(0, duration)
        self.positionSlider.setValue(1)

    def pbrate_change(self):
        '''
        재생속도 변경하는 함수
        '''
        pbrate, ok = QInputDialog.getInt(self, '재생속도 설정', '재생속도 입력 : ')

        if ok:
            if pbrate == 0:
                self.mediaPlayer.setPlaybackRate(1)
            else:
                self.mediaPlayer.setPlaybackRate(pbrate)

    def setPosition(self, position):
        """
        재생시간 변경하는 함수
        :param position: 해당 값으로 영상 시간 변경
        """
        self.mediaPlayer.setPosition(position)

    def handleError(self):
        '''
        이벤트 핸들러 오류 출력 함수
        '''
        self.playButton.setEnabled(False)
        self.error.setText("Error: " + self.mediaPlayer.errorString())

    def rightKey_pressed(self):
        '''
        오른쪽 키 눌렀을 때 작동 - 5초 앞으로 이동
        '''
        self.positionSlider.setValue(self.positionSlider.value() + 5000)
        print(self.positionSlider.value())

    def leftkey_pressed(self):
        """
        왼쪽 방향키 눌렀을 때 작동 - 5초 뒤로 이동
        """
        self.positionSlider.setValue(self.positionSlider.value() - 5000)

    def space_pressed(self):
        if self.mediaPlayer.state() == QMediaPlayer.PlayingState:
            self.mediaPlayer.pause()
        else:
            self.mediaPlayer.play()

    def keyPressEvent(self, e):
        if e.key() == Qt.Key_Right:
            self.rightKey_pressed()
        elif e.key() == Qt.Key_Left:
            self.leftkey_pressed()
        elif e.key() == Qt.Key_Space:
            self.space_pressed()

    def on_press(self, event):
        """
        마우스 클릭 인식 이벤트 핸들러
        :param event: 마울스 클릭했을 때 좌표 저장
        """

        wb = load_workbook(self.xlsx_name)
        ws = wb.worksheets[0]

        self.xpos, self.ypos = int(event.xdata), int(event.ydata)
        ch, day = vid_info(self.fileName)
        try:
            save_time = str(self.sec//60) + ':' + str(self.sec - ((self.sec//60) * 60))
        except:
            pass

        # alignment = Alignment(vertical='center', horizontal='center')
        print(self.xpos)
        ws[f'D{self.next}'].value = f'{self.xpos}, {self.ypos}'
        if self.mode == 0:
            ws[f'C{self.next}'].value = self.fileName
        ws[f'A{self.next}'].value = int(day)
        ws[f'B{self.next}'].value = int(ch)

        self.next += 1
        wb.save(self.xlsx_name)
        plt.close('all')

        if self.mode == 0:
            os.remove('original.png')

    def plt_close(self):
        pass
    def cap_img(self):
        """
        이미지 캡처 함수
        """
        try:
            os.remove('original.png')
        except:
            pass

        self.sec = self.positionSlider.value() // 1000

        os.system(fr"ffmpeg -vsync 2 -ss {self.sec} -t {self.sec+1} -i {self.fileDir} -an -vf thumbnail=25 {'./original'}.png ")

        img = plt.imread('original.png')
        plt.figure(figsize=(16, 9))
        plt.imshow(img)
        plt.connect('button_press_event', self.on_press) # 이벤트 핸들러다.....
        plt.axis('equal')
        plt.show()

    def open_image(self):
        img_dir, _ = QFileDialog.getOpenFileName(self, "Select Image", './', self.tr("Images (*.jpg *.png *.jpeg)"))
        self.mode = 1

        if img_dir != '':
            for i in range(-1, -(int(len(img_dir))), -1):
                print(i)
                if img_dir[i] == '/':
                    self.fileName = img_dir[i+1:]
                    break

            img = plt.imread(img_dir)
            plt.figure(dpi=100)
            plt.imshow(img)
            plt.connect('button_press_event', self.on_press)
            plt.axis('equal')
            plt.show()

    def img_30sec(self):
        pass

from func import excel_linecheck
app = QApplication(sys.argv)
videoplayer = VideoPlayer(n=excel_linecheck())
screen_rect = app.desktop().screenGeometry()
videoplayer.resize(screen_rect.width(), screen_rect.height())
videoplayer.show()
sys.exit(app.exec_())
